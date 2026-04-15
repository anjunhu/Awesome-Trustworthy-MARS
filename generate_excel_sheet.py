"""
Excel taxonomy — faithful to screenshots.

Row layout (matching screenshots exactly):
  Row 1: P0 blank | P1 mega-header "P1. Basic Information..." | Relevance | Rec | Generative?
  Row 2: blank    | Generative? | Output Scope | Primary goals | Application Setting | Application Domain
  Row 3: blank    | Yes | User-facing outputs | Non user-facing (intermediate) | Info access / Clinical DS / Others | Standalone vs Embedded / Single-vs-Group | domain cols
  Row 4: blank    | rotated leaf headers (Items, Expl., Reason., Conv., Media, Type | Emb., Token, Int.Retr.set, Aug.Meta., Internal Reason., Type | Type | Standalone, Embedded | Single-user, Group | Retail&E-Com., Ent./Media, Tourism/Travel, Food/Restaurant, Healthcare&Well-Being)
  Row 5: title | ID | authors | year | Keyword | doi | Assignment | Duplicates | Not-Relevant | Reason | Comments | Auxiliary Systems | Others | Modular? | [P1 cols blank in row5] | Relevance | Rec | Generative?
  Row 6+: data
"""
import re, sys
sys.path.insert(0, "/Users/anjunhu/Bookchapter")
from paper_data import KNOWN
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def fill(h): return PatternFill("solid", fgColor=h)
thin = Side(style="thin", color="BBBBBB")
med  = Side(style="medium", color="777777")
def bdr(): return Border(left=thin, right=thin, top=thin, bottom=thin)
def bdr_m(): return Border(left=med, right=med, top=med, bottom=med)
def font(bold=False, size=9, color="000000", italic=False):
    return Font(bold=bold, size=size, color=color, italic=italic)
def al(h="center", v="center", wrap=True, rot=0):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap, text_rotation=rot)

# fills
F_P0    = fill("1F3964")
F_P1    = fill("9DC3E6")
F_GEN   = fill("BDD7EE")
F_OUT_U = fill("A9D18E")
F_OUT_I = fill("E2EFDA")
F_GOAL  = fill("C5A3D0")
F_APP   = fill("F4B183")
F_DOM   = fill("FFE699")
F_REL   = fill("BDD7EE")
F_P2    = fill("2D6A9F")
F_P3    = fill("A9D18E")
F_WHITE = fill("FFFFFF")

RF_FILLS = {
    'RF0':'DDEEFF','RF1':'D6E4F0','RF2':'FCE4D6','RF3':'E2EFDA',
    'RF4':'EAD9F5','RF5':'FDEBD0','RF6':'D5F5E3','EVAL':'FFF9C4',
    'DEF':'E8F8F5','SURVEY':'F2F3F4','TBD':'FDFEFE','OTHER':'FFFFFF',
}

# ── Column layout ─────────────────────────────────────────────────────────────
# Each entry: (col_name, width, block, row5_header)
# block: P0 | GEN | OUT_U | OUT_I | GOAL | APP | DOM | REL | P2 | P3
COLS = [
    # P0 — these get their names in row 5
    ("title",            32, "P0", "title"),
    ("ID",                5, "P0", "ID"),
    ("authors",          20, "P0", "authors"),
    ("year",              6, "P0", "year"),
    ("Keyword",          16, "P0", "Keyword"),
    ("doi",              20, "P0", "doi"),
    ("Assignment",       10, "P0", "Assignment"),
    ("Duplicates",        9, "P0", "Duplicates"),
    ("Not-Relevant",      9, "P0", "Not-Relevant"),
    ("Reason",           16, "P0", "Reason"),
    ("Comments",         22, "P0", "Comments"),
    ("Auxiliary Systems",14, "P0", "Auxiliary Systems"),
    ("Others",            9, "P0", "Others"),
    ("Modular?",          8, "P0", "Modular?"),
    # P1 — Generative?
    ("Generative?",       9, "GEN", ""),
    # P1 — User-facing outputs
    ("Items",             6, "OUT_U", ""),
    ("Expl.",             6, "OUT_U", ""),
    ("Reason.",           7, "OUT_U", ""),
    ("Conv.",             6, "OUT_U", ""),
    ("Media",             6, "OUT_U", ""),
    ("Type",              8, "OUT_U", ""),
    # P1 — Non user-facing (intermediate)
    ("Emb.",              6, "OUT_I", ""),
    ("Token",             6, "OUT_I", ""),
    ("Int. Retr. set",    9, "OUT_I", ""),
    ("Aug. Meta.\n(side info)", 9, "OUT_I", ""),
    ("Internal\nReason.", 9, "OUT_I", ""),
    ("Type_i",            8, "OUT_I", ""),
    # P1 — Primary goals
    ("Info access\nType",10, "GOAL", ""),
    ("Clinical DS",       9, "GOAL", ""),
    ("Others_g",          9, "GOAL", ""),
    # P1 — Application Setting
    ("Standalone",        9, "APP", ""),
    ("Embedded",          9, "APP", ""),
    ("Single-user",       9, "APP", ""),
    ("Group",             7, "APP", ""),
    # P1 — Application Domain
    ("Retail &\nE-Com.",  9, "DOM", ""),
    ("Ent. /\nMedia",     9, "DOM", ""),
    ("Tourism /\nTravel", 9, "DOM", ""),
    ("Food /\nRestaurant",9, "DOM", ""),
    ("Healthcare &\nWell-Being",9,"DOM",""),
    # Relevance / Rec / Generative? (right side, row 1 headers)
    ("Relevance",        12, "REL", ""),
    ("Rec",               8, "REL", ""),
    ("Generative?_r",     9, "REL", ""),
    # P2 — Risk Taxonomy (our addition)
    ("Risk Family",      12, "P2", "Risk Family"),
    ("Risk Type\n(A/E)",  9, "P2", "Risk Type\n(A/E)"),
    ("Threat Tier",      10, "P2", "Threat Tier"),
    ("Topology",         10, "P2", "Topology"),
    ("Topics",           22, "P2", "Topics"),
    ("Tags",             28, "P2", "Tags"),
    # P3 — Evaluation
    ("Domain",           18, "P3", "Domain"),
    ("Metrics",          35, "P3", "Metrics"),
    ("Datasets",         30, "P3", "Datasets"),
    ("Highlights",       38, "P3", "Highlights"),
    ("Notes",            22, "P3", "Notes"),
]

def parse_readme(path="README.md"):
    text = open(path).read()
    sections = re.split(r'\n## ', text)
    rf_lookup = {
        'Foundational':'RF0','Risk Family 1':'RF1','Risk Family 2':'RF2',
        'Risk Family 3':'RF3','Risk Family 4':'RF4','Risk Family 5':'RF5',
        'Risk Family 6':'RF6','Collusion':'RF6','Fairness':'RF5',
        'Evaluation':'EVAL','Defence':'DEF','Broad Safety':'SURVEY','Uncategorised':'TBD',
    }
    papers = []
    for sec in sections:
        lines = sec.strip().split('\n')
        sec_title = lines[0].strip()
        rf = next((v for k,v in rf_lookup.items() if k in sec_title), 'OTHER')
        for title,authors,venue,arxiv,notes,tags in re.findall(
            r'\|\s*\*\*(.*?)\*\*\s*[—–-]\s*(.*?)\s*\|\s*(.*?)\s*\|\s*\[?([\d]{4}\.\d+)\]?.*?\|\s*(.*?)\s*\|\s*(.*?)\s*\|', sec):
            rfs = re.findall(r'`risk:(rf\d)`', tags)
            topics = ", ".join(re.findall(r'`topic:([^`]+)`', tags))
            risk_type = "E" if "`type:E`" in tags else ("A" if "`type:A`" in tags else "")
            papers.append(dict(
                title=title.strip(), authors=authors.strip(), year="20"+arxiv[:2],
                venue=venue.strip(), arxiv=arxiv.strip(), section=sec_title,
                rf=",".join(sorted(set(rfs))).upper() or rf,
                risk_type=risk_type, topics=topics,
                notes=notes.strip().strip('—').strip(), tags=tags.strip(),
            ))
    return papers

PAPERS = parse_readme()

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "MA-RS Taxonomy"

ncols = len(COLS)
for ci, (_, w, _, _) in enumerate(COLS):
    ws.column_dimensions[get_column_letter(ci+1)].width = w

# helper: span of a block
def span(block):
    cols = [i+1 for i,(_, _, b, _) in enumerate(COLS) if b == block]
    return cols[0], cols[-1]

def merge_write(ws, row, c1, c2, value, f, fnt, alignment):
    cell = ws.cell(row=row, column=c1, value=value)
    cell.fill = f; cell.font = fnt; cell.alignment = alignment; cell.border = bdr()
    if c2 > c1:
        ws.merge_cells(start_row=row, start_column=c1, end_row=row, end_column=c2)
        for ci in range(c1+1, c2+1):
            ws.cell(row=row, column=ci).fill = f
            ws.cell(row=row, column=ci).border = bdr()

# ── Row 1 ─────────────────────────────────────────────────────────────────────
ws.row_dimensions[1].height = 14
# colour legend in P0 area
for i,(lbl,fhex) in enumerate([("not reviewed","FFFFFF"),("already reviewed","D9EAD3"),
                                ("current reviews","FFF2CC"),("confirmed not relevant","EA9999")]):
    c = ws.cell(row=1, column=i*3+1, value=lbl)
    c.fill = fill(fhex); c.font = font(size=8, italic=True)
    c.alignment = al(h="left")

# P1 mega-header in row 1 (spans GEN..DOM)
s1,_ = span("GEN"); _,e1 = span("DOM")
merge_write(ws,1,s1,e1,"P1. Basic Information:\n(Output, Goal, Application Setting. etc.)",
            F_P1, font(bold=True,size=11), al())

# Relevance header row 1
sr,er = span("REL")
merge_write(ws,1,sr,er,"Relevance", F_REL, font(bold=True,size=10), al())

# Rec and Generative? are individual — already in REL span, just label them
# (they'll get their own cells in row 5)

# ── Row 2 ─────────────────────────────────────────────────────────────────────
ws.row_dimensions[2].height = 20
# P0 blank
s0,e0 = span("P0")
merge_write(ws,2,s0,e0,"", F_P0, font(), al())

# Generative?
sg,eg = span("GEN")
merge_write(ws,2,sg,eg,"Generative?", F_GEN, font(bold=True,size=10), al())

# Output Scope spans OUT_U + OUT_I
su,_ = span("OUT_U"); _,ei = span("OUT_I")
merge_write(ws,2,su,ei,"Output Scope", F_OUT_U, font(bold=True,size=10), al())

# Primary goals
sp,ep = span("GOAL")
merge_write(ws,2,sp,ep,"Primary goals", F_GOAL, font(bold=True,size=10), al())

# Application Setting
sa,ea = span("APP")
merge_write(ws,2,sa,ea,"Application Setting", F_APP, font(bold=True,size=10), al())

# Application Domain
sd,ed = span("DOM")
merge_write(ws,2,sd,ed,"Application Domain", F_DOM, font(bold=True,size=10), al())

# REL blank row2
merge_write(ws,2,sr,er,"", F_REL, font(), al())

# P2/P3 row2
s2,e2 = span("P2"); s3,e3 = span("P3")
merge_write(ws,2,s2,e2,"P2. Risk Taxonomy", F_P2, font(bold=True,size=10,color="FFFFFF"), al())
merge_write(ws,2,s3,e3,"P3. Evaluation",    F_P3, font(bold=True,size=10), al())

# ── Row 3 ─────────────────────────────────────────────────────────────────────
ws.row_dimensions[3].height = 20
merge_write(ws,3,s0,e0,"", F_P0, font(), al())

# "Yes" under Generative?
merge_write(ws,3,sg,eg,"Yes", F_GEN, font(bold=True,size=10), al())

# User-facing outputs
su2,eu2 = span("OUT_U")
merge_write(ws,3,su2,eu2,"User-facing outputs", F_OUT_U, font(bold=True,size=10), al())

# Non user-facing output (intermediate)
si2,ei2 = span("OUT_I")
merge_write(ws,3,si2,ei2,"Non user-facing output (intermediate)", F_OUT_I, font(bold=True,size=10), al())

# Primary goals sub-headers: "Information access" spans first col, then Clinical DS, Others
# Info access spans col 1 of GOAL only (it has a sub-row "Type" in row4)
# Clinical DS and Others are separate
goal_cols = [i+1 for i,(_, _, b, _) in enumerate(COLS) if b == "GOAL"]
# row3: "Information access" over first col, "Clinical DS" over second, "Others" over third
merge_write(ws,3,goal_cols[0],goal_cols[0],"Information access", F_GOAL, font(bold=True,size=9), al())
merge_write(ws,3,goal_cols[1],goal_cols[1],"Clinical DS",        F_GOAL, font(bold=True,size=9), al())
merge_write(ws,3,goal_cols[2],goal_cols[2],"Others",             F_GOAL, font(bold=True,size=9), al())

# Application Setting sub-headers
app_cols = [i+1 for i,(_, _, b, _) in enumerate(COLS) if b == "APP"]
merge_write(ws,3,app_cols[0],app_cols[1],"Standalone vs. Embedded", F_APP, font(bold=True,size=9), al())
merge_write(ws,3,app_cols[2],app_cols[3],"Single-vs. Group",        F_APP, font(bold=True,size=9), al())

# Application Domain — blank row3 (leaf headers in row4)
merge_write(ws,3,sd,ed,"", F_DOM, font(), al())

# REL blank
merge_write(ws,3,sr,er,"", F_REL, font(), al())
merge_write(ws,3,s2,e2,"", F_P2, font(), al())
merge_write(ws,3,s3,e3,"", F_P3, font(), al())

# ── Row 4: rotated leaf headers for P1; blank for P0; col names for P2/P3 ─────
ws.row_dimensions[4].height = 75
merge_write(ws,4,s0,e0,"", F_P0, font(), al())

BLOCK_FILL4 = {"GEN":F_GEN,"OUT_U":F_OUT_U,"OUT_I":F_OUT_I,
               "GOAL":F_GOAL,"APP":F_APP,"DOM":F_DOM,
               "REL":F_REL,"P2":F_P2,"P3":F_P3}
BLOCK_FC4   = {"P2":"FFFFFF"}

for ci,(name,_,block,_) in enumerate(COLS):
    if block == "P0": continue
    c = ws.cell(row=4, column=ci+1)
    # display name: strip internal suffix used to deduplicate
    display = name.replace("_i","").replace("_g","").replace("_r","")
    c.value = display
    c.fill  = BLOCK_FILL4.get(block, F_WHITE)
    c.font  = font(bold=True, size=9, color=BLOCK_FC4.get(block,"000000"))
    c.alignment = Alignment(horizontal="center", vertical="bottom",
                            wrap_text=False, text_rotation=45)
    c.border = bdr()

# ── Row 5: P0 column name headers + P2/P3 repeated; P1 blank ─────────────────
ws.row_dimensions[5].height = 22
for ci,(name,_,block,row5hdr) in enumerate(COLS):
    c = ws.cell(row=5, column=ci+1)
    c.value = row5hdr
    if block == "P0":
        c.fill = F_P0; c.font = font(bold=True,size=10,color="FFFFFF")
        c.alignment = al(h="center",v="center")
    elif block in ("P2","P3"):
        c.fill = BLOCK_FILL4[block]
        c.font = font(bold=True,size=9,color=BLOCK_FC4.get(block,"000000"))
        c.alignment = al(h="center",v="center")
    else:
        c.fill = BLOCK_FILL4.get(block, F_WHITE)
        c.font = font(size=9)
        c.alignment = al()
    c.border = bdr()

ws.freeze_panes = "A6"
ws.auto_filter.ref = f"A5:{get_column_letter(ncols)}5"

# ── col index by name ─────────────────────────────────────────────────────────
COL_IDX = {name: i+1 for i,(name,_,_,_) in enumerate(COLS)}

def infer_tier(tags):
    if any(x in tags for x in ("rf1","rf2")): return "Compromise"
    if "rf3" in tags: return "Compromise"
    if "rf4" in tags: return "Misalignment"
    if "rf5" in tags or "rf6" in tags: return "Drift/Misalignment"
    return ""

def sc(row, name, value, rf):
    idx = COL_IDX.get(name)
    if idx is None: return
    c = ws.cell(row=row, column=idx, value=value)
    c.fill = fill(RF_FILLS.get(rf,'FFFFFF'))
    c.font = font(size=9)
    c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    c.border = bdr()

# ── Data rows ─────────────────────────────────────────────────────────────────
for ri, p in enumerate(PAPERS):
    row = ri + 6
    ws.row_dimensions[row].height = 55
    rf = p['rf'].split(',')[0]
    k  = KNOWN.get(p['arxiv'], {})

    # fill all cells
    for ci in range(ncols):
        c = ws.cell(row=row, column=ci+1)
        c.fill = fill(RF_FILLS.get(rf,'FFFFFF'))
        c.font = font(size=9)
        c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        c.border = bdr()

    sc(row,"title",    p['title'],   rf)
    sc(row,"authors",  p['authors'], rf)
    sc(row,"year",     p['year'],    rf)
    sc(row,"Keyword",  p['topics'],  rf)
    sc(row,"doi",      f"https://arxiv.org/abs/{p['arxiv']}", rf)
    sc(row,"Relevance",p['section'], rf)
    sc(row,"Risk Family", p['rf'],   rf)
    sc(row,"Risk Type\n(A/E)", p['risk_type'], rf)
    sc(row,"Threat Tier", infer_tier(p['tags']), rf)
    sc(row,"Topology", next(iter(re.findall(r'`topo:([^`]+)`', p['tags'])), ""), rf)
    sc(row,"Topics",   p['topics'],  rf)
    sc(row,"Tags",     p['tags'],    rf)
    sc(row,"Notes",    p['notes'],   rf)

    domain = k.get('domain','')
    sc(row,"Domain",    domain, rf)
    sc(row,"Metrics",   "\n".join(k.get('metrics',[])),    rf)
    sc(row,"Datasets",  "\n".join(k.get('datasets',[])),   rf)
    sc(row,"Highlights","\n".join(k.get('highlights',[])), rf)

    # Generative? checkbox
    if any(x in domain.lower() for x in ("recommendation","recsys","conversational","generative")):
        sc(row,"Generative?", "✓", rf)
        sc(row,"Generative?_r","✓", rf)

    # Application domain checkboxes
    d = domain.lower()
    if "recommendation" in d or "recsys" in d:
        sc(row,"Retail &\nE-Com.", "✓", rf)
    if "healthcare" in d:
        sc(row,"Healthcare &\nWell-Being", "✓", rf)

    # Application setting
    sc(row,"Standalone", "✓", rf)
    if "group" in p['tags']:
        sc(row,"Group", "✓", rf)
    else:
        sc(row,"Single-user", "✓", rf)

wb.save("mars_taxonomy.xlsx")
print(f"Saved → mars_taxonomy.xlsx  ({len(PAPERS)} papers)")
